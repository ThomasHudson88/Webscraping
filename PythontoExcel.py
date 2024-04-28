import openpyxl as xl
from openpyxl.styles import Font

#Now we are going to take date from python and put it into excel
#Create a new Excel doc
wb = xl.Workbook()

#This is making Sheet 1 the active sheet
ws = wb.active
ws.title = 'First Sheet'

wb.create_sheet(index=1,title='Second Sheet')

#Write content to the sheet
ws['A1'] = 'Invoice'
ws['A1'].font = Font('Times New Roman',size= 24,italic=False,bold=True)

headerfont = Font('Times New Roman',size= 24,italic=False,bold=True)

ws['A2'] = 'Tires'
ws['A3'] = 'Brakes'
ws['A4'] = 'Alignment'

ws['B2'] = 450
ws['B3'] = 225
ws['B4'] = 150

#merge cells A1 and B1
ws.merge_cells('A1:B1')

#Can unmerge with 'ws.unmerge_cells('A1:B1')'

ws['A8'] = 'Total'
ws['A8'].font = Font(size=16,bold=True)

#Now we will input a formula
ws['B8'] = '=sum(B2:B7)'

#Change the dimensions, this will set the width the 20 pixels
ws.column_dimensions['A'].width = 20

#Now we are going to take one excel and put it into our new excel here
write_sheet = wb['Second Sheet']
read_wb = xl.load_workbook("ProduceReport.xlsx")
read_ws = read_wb['ProduceReport']

#Now we can create a loop to read from one and write to another
for row in read_ws:
    #using list comprehension we create a list with all of the values in it
    my_list = [i.value for i in row]
    write_sheet.append(my_list)

#get the max row
max_row = write_sheet.max_row

#Add the total and average formulas
write_sheet.cell(max_row+2,1).value = "Totals:"
write_sheet.cell(max_row+3,1).value = "Averages:"

#SUM formula
write_sheet.cell(max_row+2,3).value = f"=SUM(C1:C{max_row})"
write_sheet.cell(max_row+2,4).value = f"=SUM(D1:D{max_row})"

#AVERAGE forumla
write_sheet.cell(max_row+3,3).value = f"=AVERAGE(C1:C{max_row})"
write_sheet.cell(max_row+3,4).value = f"=AVERAGE(D1:D{max_row})"

write_sheet.column_dimensions['A'].width = 15
write_sheet.column_dimensions['D'].width = 15

#Format the numbers correctly
for cell in write_sheet["C:C"]:
    cell.number_format = '#,##0'
for cell in write_sheet["D:D"]:
    cell.number_format = u'"$"#,##0.00'

#Save it
wb.save('PythontoExcel.xlsx')