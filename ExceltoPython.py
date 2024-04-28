import openpyxl as xl

#LOAD the workbook and sheet names
wb = xl.load_workbook("example.xlsx")
sn = wb.sheetnames

print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']
'''
print(sheet1)
print(cellA1)

print(cellA1.value)

print(cellA1.row)
print(cellA1.column)

print(cellA1.coordinate)

#can use this cell function to findout the vlaue of a cell , this is for cell B2
print(sheet1.cell(1,2).value)

#this method can show how many rows and columns have data
print(sheet1.max_row)
print(sheet1.max_column)
'''

#This will print all of the Fruits in Column B
#The column is not going to change, we want to iterate over the rows
for row in range(1,sheet1.max_row+1):
    print(sheet1.cell(row,2).value)

print(xl.utils.get_column_letter(900))
print(xl.utils.column_index_from_string('AHP'))

#This is iterating through the entire sheet row by row and then cell by cell
for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

#iter_rows takes 3 arguments
for currentrow in sheet1.iter_rows(min_row=1,max_row=sheet1.max_row,max_col=sheet1.max_column):
    #print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)
#this prints out each row in the sheet line by line for the frist 3 cells, that is date, fruit, count in the data