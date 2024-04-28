
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##
#We can think of this ogject table_rows as an iterable list that we can get info from
table_rows = soup.findAll("tr")

for row in table_rows[1:6]:
    #We are grabbing the first 5 rows of the table and now we need to get the 'TD' tags for the info
    #print(row.text)
   
     td = row.findAll("td"),
     no = td[0].text,
     movie = td[1].text,
     release_date = td[8].text,
     theatres = td[6].text,
     total_gross = td[5].text

#Now we are going to create the Excel and also format the sheet
wb = xl.Workbook()

#This is making Sheet 1 the active sheet
ws = wb.active
ws.title = 'Box Office Report'
#Format the width
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 30

#Now we make the headers
headerfont = Font('Times New Roman',size= 14,italic=False,bold=True)
ws['A1'] = 'No.'
ws['B1'] = 'Realease Date'
ws['C1'] = 'Number of Theaters'
ws['D1'] = 'Total Gross'
ws['E1'] = 'Average Gross by Theater'

ws['A1'].font = headerfont
ws['B1'].font = headerfont
ws['C1'].font = headerfont
ws['D1'].font = headerfont
ws['E1'].font = headerfont









#Save it
wb.save('WebScrapeMovies.xlsx')